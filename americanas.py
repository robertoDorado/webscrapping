# _*_ coding: utf-8 _*_
from selenium import webdriver
from bs4 import BeautifulSoup
import openpyxl
import os

search = input('qual será o produto a ser procurado: ')

if search.isalpha() == False:
    print('valor de busca inválido')
    exit()

try:
    offset = float(input('qual será a página a ser procurada: '))
except:
    print('valor inválido')
    exit()

page_value = str((offset - 1) * 24)

page = '?limit=24&offset=' + page_value
url = 'https://www.americanas.com.br/busca/' + search + page

navigation = webdriver.Chrome('./chromedriver-win64/chromedriver.exe')
navigation.get(url)

navigation.implicitly_wait(10)
site = BeautifulSoup(navigation.page_source, 'html.parser')
products = site.find_all('a', {'class', 'JOEpk'})

product_data = []

for product in products:
    
    if product.find('h3', {'class', 'gUjFDF'}):
        title = product.find('h3', {'class', 'gUjFDF'})
    else:
        continue
    
    if product.find('span', {'class', 'liXDNM'}) != None:
        price = product.find('span', {'class', 'liXDNM'})
    else:
        continue
    
    if product.find('span', {'class', 'JbUli'}) != None:
        old_price = product.find('span', {'class', 'JbUli'})
    else:
        continue
    
    current_price = price.text.replace('R$', '')
    old_price_data = old_price.text.replace('R$', '')
    
    try:
        old_price_value = float(old_price_data.replace(".", "").replace(",", ".")) if len(old_price_data) > 7 else float(old_price_data.replace(",", "."))
        current_price = float(current_price.replace(".", "").replace(",", ".")) if len(current_price) > 7 else float(current_price.replace(",", "."))
        percent_change = float(format((current_price - old_price_value) / old_price_value, '.2f'))
    except:
        continue
    
    product_data.append({
        'name': title.text,
        'current_price': price.text,
        'old_price': old_price.text,
        'percent_change': str(percent_change * 100) + "%",
        'current_price_value': current_price
    })

product_data = sorted(product_data, key=lambda product: product['current_price_value'], reverse=True)
excel_file = f"{search}-americanas.xlsx"

if os.path.exists(excel_file) == False:
    
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "produtos-americanas"
    
    worksheet["A1"] = "Nome do produto"
    worksheet["B1"] = "Preço Atual"
    worksheet["C1"] = "Preço Anterior"
    worksheet["D1"] = "Variação percentual"
        
else:
    workbook = openpyxl.load_workbook(excel_file)
    worksheet = workbook.active

next_index = worksheet.max_row + 1

for product in product_data:
    worksheet[f"A{next_index}"] = product["name"]
    worksheet[f"B{next_index}"] = product["current_price"]
    worksheet[f"C{next_index}"] = product["old_price"]
    worksheet[f"D{next_index}"] = product["percent_change"]
    next_index += 1

workbook.save(excel_file)    
print(f'produtos salvo com sucesso em {excel_file}')
navigation.quit()